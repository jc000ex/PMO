"""一次性脚本：将Excel项目数据转为projects.json"""
import openpyxl, json, os

src = os.path.dirname(__file__)
files = [f for f in os.listdir(src) if f.endswith('.xlsx')]
if not files:
    print("没找到Excel文件")
    exit()

wb = openpyxl.load_workbook(os.path.join(src, files[0]))
ws = wb[wb.sheetnames[0]]

projects = []
for row in ws.iter_rows(min_row=2, values_only=True):
    name = (row[0] or '').strip()
    if not name:
        continue
    projects.append({
        "id": (row[3] or '').strip(),
        "name": name,
        "phase": (row[1] or '').strip(),
        "status": (row[2] or '').strip(),
        "priority": (row[4] or '').strip(),
        "customer": (row[5] or '').strip(),
        "description": (row[6] or '').strip(),
        "pm": (row[7] or '').strip(),
        "startDate": str(row[8])[:10] if row[8] else '',
        "endDate": str(row[9])[:10] if row[9] else '',
        "actualEndDate": str(row[10])[:10] if row[10] else '',
        "members": (row[11] or '').strip(),
        "businessPhase": False,
        "businessProgress": '',
        "currentFocus": '',
        "risks": [],
        "milestones": [],
        "updates": []
    })

out = os.path.join(src, 'data', 'projects.json')
json.dump(projects, open(out, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
print(f'转换完成：{len(projects)} 个项目 -> {out}')
